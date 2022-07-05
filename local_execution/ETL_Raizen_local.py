import pandas as pd
from openpyxl import load_workbook
import os
import datetime
from datetime import datetime

#locale.setlocale(locale.LC_ALL, 'pt_PT.UTF-8')

#Download of the raw_data

def download_file_xls():
    os.system('mkdir raw_data')
    from urllib import request
    file_url = 'https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls'
    file = './raw_data/vendas-combustiveis-m3.xls'
    request.urlretrieve(file_url , file)

#convert file xls in xlsx

def convert_xls_in_xlsx():
    os.system('''libreoffice --headless --invisible --convert-to xlsx ./raw_data/vendas-combustiveis-m3.xls --outdir ./raw_data/''')

path = "./raw_data/vendas-combustiveis-m3.xlsx"
sheets = {"DPCache_m3": "oil_derivative", "DPCache_m3_2": "diesel"}

#extract tables oil_derivative and diesel

def extract_tables_file():

    def extract_xlsx_sheet(file_location, sheet, output_filelocation):
        wb = load_workbook(file_location)
        sheets = wb.sheetnames

        for s in sheets:
            if s != sheet:
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
        
        wb.save(output_filelocation)
    
    for i in sheets:
        extract_xlsx_sheet(path, i , f"{sheets[i]}.xlsx")

#Transform DataFrame and Saving in parquet format

os.system('pip install fastparquet')

def transform_n_load_datas():
    for i in sheets:
        
        # read raw data

        df = pd.read_excel(f"{sheets[i]}.xlsx")

        # Process column names

        df['product'] = df['COMBUSTÍVEL'].str.split(' \(').str[0]

        df['unit'] = df['COMBUSTÍVEL'].str.split(" \(").str[1].str.split("\)").str[0]

        df = df.rename(columns={'ESTADO': 'uf'})

        months_dict = {}
        for j in df.columns:
            try:
                month_number = datetime.datetime.strptime(j, "%b").strftime('%m')
                months_dict[j] = month_number
            except:
                pass
        for j in months_dict:
            df = df.rename(columns={j: months_dict[i]})
            
        # drop useless columns

        df = df.drop(['COMBUSTÍVEL', 'REGIÃO', 'TOTAL'],axis=1)

        # Unpivot months

        df =  pd.melt(df, id_vars=['product', 'uf', 'ANO', 'unit'], var_name=['month'], value_name='volume')

        # Create year-month column

        df['year_month'] = df['ANO'].astype(str) + '-' + df['month'].astype(str)

        # Create timestamp column

        df['created_at'] = datetime.today().replace(microsecond=0)

        # Drop not asked columns

        df = df.drop(['ANO', 'month'], axis=1)

        # Fill nans with 0

        df = df.fillna(0)

        # Format dtypes

        df['product'] = df['product'].astype(str)
        df['uf'] = df['uf'].astype(str)
        df['unit'] = df['unit'].astype(str)
        df['volume'] = df['volume'].astype(float)
        df['year_month'] = df['year_month'].astype(str)
        df['created_at'] = pd.to_datetime(df['created_at'])

        #saving files in parquet format by product and year-months

        persist_path = f'{sheets[i]}'
        df.to_parquet(persist_path, engine='fastparquet', partition_cols=['product', 'year_month'])

download_file_xls()
convert_xls_in_xlsx()
extract_tables_file()
transform_n_load_datas()

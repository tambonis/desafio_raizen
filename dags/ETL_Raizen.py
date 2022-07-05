#!/usr/bin/env python
# coding: utf-8

# In[1]:


#####################################################################################################################
#ANP Fuel Sales ETL Test - raízen
# Goal
#The developed pipeline is meant to extract and structure the underlying data of two of these tables:#

#    Sales of oil derivative fuels by UF and product
#    Sales of diesel by UF and type

#The totals of the extracted data must be equal to the totals of the pivot tables.
# Tiago Tambonis - 30/06 - ttamboniss@gmail.com
#####################################################################################################################


# In[2]:


from airflow import DAG
from airflow.operators.python import PythonOperator
import pandas as pd
from openpyxl import load_workbook
import os
import datetime
from urllib import request
import warnings
import pip

with warnings.catch_warnings():
    warnings.filterwarnings("ignore",category=DeprecationWarning)
    
path = "./dados_extracao/vendas-combustiveis-m3.xlsx"
sheets = {"DPCache_m3": "oil_derivative", "DPCache_m3_2": "diesel"}


# In[3]:


def import_or_install(package):
    try:
        __import__(package)
        print(f'Package {package} installed.')
    except ImportError:
        pip.main(['package', package])


# In[4]:


def get_file():
    os.system('mkdir dados_extracao')
    file_url = 'https://github.com/raizen-analytics/data-engineering-test/raw/master/assets/vendas-combustiveis-m3.xls'
    file = './dados_extracao/vendas-combustiveis-m3.xls'
    request.urlretrieve(file_url , file)    


# In[5]:


def xls_to_xlsx():
    os.system('''libreoffice --headless --invisible --convert-to xlsx ./dados_extracao/vendas-combustiveis-m3.xls --outdir ./dados_extracao/''')


# In[6]:


#De uma ajustada nessa, está copiada idêntica
def extract_tables_file():

    def extract_xlsx_sheet(file_location, sheet, output_filelocation):
        
        wb = load_workbook(file_location)
        sheets = wb.sheetnames

        for s in sheets:
            if s != sheet:
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
        
        wb.save(output_filelocation)
    
    for i in sheets:
        extract_xlsx_sheet(path, i , f"{sheets[i]}.xlsx")


# In[7]:


def fn(month_name):
    month_name = datetime.datetime.strptime(month_name, '%b').month
    return(month_name)


# In[8]:


def data_quality(dq, name):

    dq['year_month_dt'] = pd.to_datetime(dq['year_month']).dt.strftime('%Y-%m')

    last_month = pd.to_datetime(dq['year_month_dt'].sort_values(ascending=True).unique()[-2]).strftime('%Y-%m')

    current_month = pd.to_datetime(dq['year_month_dt'].sort_values(ascending=True).unique()[-1]).strftime('%Y-%m')

    std_last = dq[dq['year_month_dt']==last_month]['volume'].std()
    mean_last = dq[dq['year_month_dt']==last_month]['volume'].mean()

    std_cur = dq[dq['year_month_dt']==current_month]['volume'].std()
    mean_cur = dq[dq['year_month_dt']==current_month]['volume'].mean()

    print(f"{name}")
    print('\n')
    print('###########################')
    print(f'Last month measures')
    print(f'Deviation: {std_last}.')
    print(f'Mean: {mean_last}')
    print('###########################')
    print(f'Current month measures')
    print(f'Deviation: {std_cur}.')
    print(f'Mean: {mean_cur}')
    print('###########################')
    print('Relative differences: ')
    print(f'Deviation: {std_cur/std_last}.')
    print(f'Mean: {mean_cur/mean_last}.')
    print('###########################')
    
    if (abs(mean_cur-mean_last)>((mean_cur+mean_last)/2)*0.10):
        print('Possible data quality issue.')
    else: print('Mean data quality ok.')

    if (abs(std_cur-std_last)>((std_cur+std_last)/2)*0.10):
        print('Possible data quality issue.')
    else: print('Std data quality ok.')


# In[9]:


def get_transform():

    for i in sheets:

        df = pd.read_excel(f"{sheets[i]}.xlsx")

        df['product'] = df['COMBUSTÍVEL'].str.split(' \(').str[0]
        df['unit'] = df['COMBUSTÍVEL'].str.split(" \(").str[1].str.split("\)").str[0]
        df = df.rename(columns={'ESTADO': 'uf'}) #posso remover

        df = df.drop(['COMBUSTÍVEL', 'REGIÃO', 'TOTAL'],axis=1)

        df =  pd.melt(df, id_vars=['product', 'uf', 'ANO', 'unit'], 
                      var_name=['month'], value_name='volume')

        df['month'] = df['month'].str.lower()

        months={'jan':'jan', 'fev':'feb', 'abr':'apr', 'mai':'may', 
                'ago':'aug', 'set':'sep', 'out':'oct', 'dez':'dec'}

        df['month'] = df.month.replace(months)

        df["month"] = df["month"].apply(fn)

        df.rename(columns={'ANO':'ano'}, inplace=True)

        df['year_month'] = df['ano'].astype(str) + '-' + df['month'].astype(str)
        df['created_at'] = datetime.datetime.today().replace(microsecond=0)
        df = df.drop(['ano', 'month'], axis=1)
        df = df.fillna(0)

        df['product'] = df['product'].astype(str)
        df['uf'] = df['uf'].astype(str)
        df['unit'] = df['unit'].astype(str)
        df['volume'] = df['volume'].astype(float)
        df['year_month'] = df['year_month'].astype(str)
        df['created_at'] = pd.to_datetime(df['created_at'])    

        data_quality(df, f"{sheets[i]}")

        persist_path = f'{sheets[i]}'
        df.to_parquet(persist_path, engine='fastparquet', partition_cols=['product', 'year_month'])

        print('\n')
        print('Visual check: ')
        print(df.head())


# In[10]:

if False: import_or_install('fastparquet')

if False: #Local?    
    get_file()
    xls_to_xlsx()
    extract_tables_file()
    get_transform()


# In[ ]:


with DAG('ETL_ANP_Raízen', start_date = datetime.datetime(2022,7,3),
           schedule_interval = '30 * * * *' , catchup = False) as dag:

    get_file = PythonOperator(
         task_id = 'get_file',
         python_callable = get_file      
    )
    
    xls_to_xlsx = PythonOperator(
         task_id = 'xls_to_xlsx',
         python_callable = xls_to_xlsx       
    )

    extract_tables_file = PythonOperator(
         task_id = 'extract_tables_file',
         python_callable = extract_tables_file
    )

    get_transform = PythonOperator(
         task_id = 'get_transform',
         python_callable = get_transform
    )

    get_file >> xls_to_xlsx >> extract_tables_file >> get_transform

